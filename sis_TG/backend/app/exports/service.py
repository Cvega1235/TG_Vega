import csv
import io
from datetime import datetime

from sqlalchemy import func, or_
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable,
)
from reportlab.graphics.shapes import Drawing, Rect, String

from app.restaurants.models import Restaurant, RestaurantScore, RestaurantMLScore


# ---------------------------------------------------------------------------
# Helpers de datos
# ---------------------------------------------------------------------------

def _get_top_scored(db: Session, limit: int = 50, **filters) -> list[dict]:
    """Retorna los restaurantes mejor puntuados, ordenados por score desc."""
    query = (
        db.query(Restaurant, RestaurantScore.total_score)
        .join(RestaurantScore)
    )
    if filters.get("fuente"):
        query = query.filter(Restaurant.fuente == filters["fuente"])
    if filters.get("zona"):
        query = query.filter(Restaurant.zona == filters["zona"])
    if filters.get("status"):
        query = query.filter(Restaurant.status == filters["status"])
    if filters.get("rating_min") is not None:
        query = query.filter(Restaurant.rating >= filters["rating_min"])
    if filters.get("rating_max") is not None:
        query = query.filter(Restaurant.rating <= filters["rating_max"])
    if filters.get("search"):
        query = query.filter(
            or_(
                Restaurant.nombre.ilike(f"%{filters['search']}%"),
                Restaurant.direccion.ilike(f"%{filters['search']}%"),
            )
        )

    query = query.order_by(RestaurantScore.total_score.desc()).limit(limit)
    rows = []
    for r, score in query.all():
        rows.append({
            "nombre": r.nombre,
            "fuente": r.fuente,
            "zona": r.zona or "",
            "direccion": r.direccion or "",
            "telefono": r.telefono or "",
            "rating": float(r.rating) if r.rating else "",
            "num_resenas": r.num_resenas or 0,
            "tipo_cocina": r.tipo_cocina or "",
            "precio": r.precio or "",
            "status": r.status,
            "score": float(score) if score else "",
            "tiene_embutidos": "Si" if r.tiene_embutidos else ("No" if r.tiene_embutidos is False else "-"),
        })
    return rows


def _get_summary_stats(db: Session) -> dict:
    """Retorna estadisticas globales para el resumen del reporte."""
    total = db.query(func.count(Restaurant.id)).scalar() or 0
    avg_rating = db.query(func.avg(Restaurant.rating)).filter(
        Restaurant.rating.isnot(None)
    ).scalar()
    high_affinity = db.query(func.count(RestaurantMLScore.id)).filter(
        RestaurantMLScore.composite_score >= 70
    ).scalar() or 0
    clients = db.query(func.count(Restaurant.id)).filter(
        Restaurant.status == "cliente"
    ).scalar() or 0
    with_embutidos = db.query(func.count(Restaurant.id)).filter(
        Restaurant.tiene_embutidos == True  # noqa: E712
    ).scalar() or 0

    zone_rows = (
        db.query(Restaurant.zona, func.count(Restaurant.id))
        .filter(Restaurant.zona.isnot(None))
        .group_by(Restaurant.zona)
        .order_by(func.count(Restaurant.id).desc())
        .limit(8)
        .all()
    )
    cuisine_rows = (
        db.query(Restaurant.tipo_cocina, func.count(Restaurant.id))
        .filter(Restaurant.tipo_cocina.isnot(None), Restaurant.tipo_cocina != "")
        .group_by(Restaurant.tipo_cocina)
        .order_by(func.count(Restaurant.id).desc())
        .limit(8)
        .all()
    )
    status_rows = (
        db.query(Restaurant.status, func.count(Restaurant.id))
        .group_by(Restaurant.status)
        .all()
    )

    return {
        "total": total,
        "avg_rating": round(float(avg_rating), 2) if avg_rating else None,
        "high_affinity": high_affinity,
        "clients": clients,
        "with_embutidos": with_embutidos,
        "by_zone": [(r[0], r[1]) for r in zone_rows],
        "by_cuisine": [(r[0], r[1]) for r in cuisine_rows],
        "by_status": {r[0]: r[1] for r in status_rows},
    }


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------

EXPORT_COLUMNS = [
    "nombre", "fuente", "zona", "direccion", "telefono", "rating",
    "num_resenas", "tipo_cocina", "precio", "status", "score", "tiene_embutidos",
]
COLUMN_HEADERS = [
    "Nombre", "Fuente", "Zona", "Direccion", "Telefono", "Rating",
    "Resenas", "Tipo Cocina", "Precio", "Estado", "Score", "Embutidos",
]


def export_csv(db: Session, **filters) -> str:
    rows = _get_top_scored(db, limit=200, **filters)
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=EXPORT_COLUMNS)
    writer.writerow(dict(zip(EXPORT_COLUMNS, COLUMN_HEADERS)))
    for row in rows:
        writer.writerow({k: row.get(k, "") for k in EXPORT_COLUMNS})
    return output.getvalue()


# ---------------------------------------------------------------------------
# Excel con grafico y hoja de resumen
# ---------------------------------------------------------------------------

def export_excel(db: Session, **filters) -> bytes:
    rows = _get_top_scored(db, limit=50, **filters)
    stats = _get_summary_stats(db)
    wb = Workbook()

    # --- Hoja 1: Resumen ---
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"

    BLUE = "2563EB"
    LIGHT_BLUE = "DBEAFE"
    GREEN = "16A34A"
    LIGHT_GREEN = "DCFCE7"

    h_font = Font(bold=True, color="FFFFFF", size=11)
    h_fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ws_resumen.merge_cells("A1:D1")
    title_cell = ws_resumen["A1"]
    title_cell.value = f"Reporte de Inteligencia de Mercado - Don Piotr  |  {datetime.now().strftime('%d/%m/%Y')}"
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center")
    ws_resumen.row_dimensions[1].height = 28

    # KPIs
    kpis = [
        ("Total Restaurantes", stats["total"], BLUE, LIGHT_BLUE),
        ("Alta Afinidad (Score>=70)", stats["high_affinity"], GREEN, LIGHT_GREEN),
        ("Clientes Actuales", stats["clients"], "7C3AED", "EDE9FE"),
        ("Con Embutidos en Menu", stats["with_embutidos"], "EA580C", "FED7AA"),
        ("Rating Promedio", stats["avg_rating"] or "N/A", "CA8A04", "FEF9C3"),
    ]
    ws_resumen.append([])
    ws_resumen.append(["INDICADORES CLAVE"])
    ws_resumen["A3"].font = Font(bold=True, size=11)

    for i, (label, value, fg, bg) in enumerate(kpis, start=4):
        ws_resumen.cell(row=i, column=1, value=label).font = Font(bold=True)
        val_cell = ws_resumen.cell(row=i, column=2, value=value)
        val_cell.font = Font(bold=True, color=fg, size=12)
        val_cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        val_cell.alignment = Alignment(horizontal="center")

    # Distribucion por zona
    ws_resumen.append([])
    row_z = len(kpis) + 6
    ws_resumen.cell(row=row_z, column=1, value="DISTRIBUCION POR ZONA").font = Font(bold=True, size=11)
    ws_resumen.cell(row=row_z + 1, column=1, value="Zona").font = Font(bold=True, color="FFFFFF")
    ws_resumen.cell(row=row_z + 1, column=1).fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
    ws_resumen.cell(row=row_z + 1, column=2, value="Restaurantes").font = Font(bold=True, color="FFFFFF")
    ws_resumen.cell(row=row_z + 1, column=2).fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")

    zone_start = row_z + 2
    for j, (zona, count) in enumerate(stats["by_zone"]):
        fill_color = LIGHT_BLUE if j % 2 == 0 else "FFFFFF"
        ws_resumen.cell(row=zone_start + j, column=1, value=zona).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        ws_resumen.cell(row=zone_start + j, column=2, value=count).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    # Grafico de barras por zona
    if stats["by_zone"]:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Restaurantes por Zona"
        chart.y_axis.title = "Cantidad"
        chart.style = 10
        chart.width = 14
        chart.height = 10

        data_ref = Reference(ws_resumen, min_col=2, min_row=row_z + 1, max_row=zone_start + len(stats["by_zone"]) - 1)
        cats_ref = Reference(ws_resumen, min_col=1, min_row=zone_start, max_row=zone_start + len(stats["by_zone"]) - 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws_resumen.add_chart(chart, "D4")

    ws_resumen.column_dimensions["A"].width = 28
    ws_resumen.column_dimensions["B"].width = 16

    # --- Hoja 2: Top 50 Prospectos ---
    ws_top = wb.create_sheet("Top 50 Prospectos")

    ws_top.merge_cells("A1:I1")
    top_title = ws_top["A1"]
    top_title.value = "Top 50 Prospectos Mejor Puntuados - Don Piotr"
    top_title.font = Font(bold=True, size=13, color="FFFFFF")
    top_title.fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
    top_title.alignment = Alignment(horizontal="center")
    ws_top.row_dimensions[1].height = 24

    top_headers = ["#", "Nombre", "Zona", "Tipo Cocina", "Rating", "Resenas", "Estado", "Score", "Embutidos"]
    top_keys = ["nombre", "zona", "tipo_cocina", "rating", "num_resenas", "status", "score", "tiene_embutidos"]

    for col, h in enumerate(top_headers, 1):
        cell = ws_top.cell(row=2, column=col, value=h)
        cell.font = h_font
        cell.fill = h_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    for i, row in enumerate(rows, 1):
        fill_color = LIGHT_BLUE if i % 2 == 0 else "FFFFFF"
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        ws_top.cell(row=i + 2, column=1, value=i).fill = row_fill
        for col, key in enumerate(top_keys, 2):
            c = ws_top.cell(row=i + 2, column=col, value=row.get(key, ""))
            c.fill = row_fill
            c.border = thin
            if key == "score":
                c.font = Font(bold=True)
                c.alignment = Alignment(horizontal="center")

    col_widths = [4, 30, 16, 18, 8, 10, 14, 10, 12]
    for col, w in enumerate(col_widths, 1):
        ws_top.column_dimensions[chr(64 + col)].width = w

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ---------------------------------------------------------------------------
# PDF — colores del sistema
# ---------------------------------------------------------------------------

PRIMARY_HEX = colors.HexColor("#9B1C2E")
PRIMARY_LIGHT_HEX = colors.HexColor("#FEE2E2")
PRIMARY_DARK_HEX = colors.HexColor("#7F1D1D")
GREEN_HEX = colors.HexColor("#16A34A")
GREEN_LIGHT_HEX = colors.HexColor("#DCFCE7")
YELLOW_HEX = colors.HexColor("#EAB308")
GRAY_HEX = colors.HexColor("#6B7280")
GRAY_LIGHT_HEX = colors.HexColor("#F9FAFB")
BORDER_HEX = colors.HexColor("#E5E7EB")
TEXT_DARK = colors.HexColor("#111827")
TEXT_MID = colors.HexColor("#374151")

# Backwards compat para Excel (no cambia)
BLUE_HEX = colors.HexColor("#2563EB")
LIGHT_BLUE_HEX = colors.HexColor("#DBEAFE")
ORANGE_HEX = colors.HexColor("#EA580C")


# ---------------------------------------------------------------------------
# Helpers PDF
# ---------------------------------------------------------------------------

def _get_top_prospects_pdf(db: Session, limit: int = 3) -> list[dict]:
    rows = (
        db.query(
            Restaurant.nombre, Restaurant.zona, Restaurant.tipo_cocina,
            Restaurant.rating, Restaurant.status, Restaurant.tiene_embutidos,
            RestaurantScore.total_score,
            RestaurantScore.cuisine_score,
            RestaurantScore.rating_score,
            RestaurantScore.reviews_score,
            RestaurantScore.zone_score,
        )
        .join(RestaurantScore)
        .filter(Restaurant.status.notin_(["cliente", "no_interesado"]))
        .order_by(RestaurantScore.total_score.desc())
        .limit(limit)
        .all()
    )
    return [
        {
            "nombre": r[0], "zona": r[1], "tipo_cocina": r[2],
            "rating": float(r[3]) if r[3] else None,
            "status": r[4], "tiene_embutidos": r[5],
            "total_score": float(r[6]),
            "cuisine_score": float(r[7]) if r[7] else None,
            "rating_score": float(r[8]) if r[8] else None,
            "reviews_score": float(r[9]) if r[9] else None,
            "zone_score": float(r[10]) if r[10] else None,
        }
        for r in rows
    ]


def _get_reasons(prospect: dict) -> list[str]:
    reasons = []
    if (prospect.get("cuisine_score") or 0) >= 20:
        reasons.append("Alta afinidad de productos con el catalogo Don Piotr")
    if (prospect.get("rating_score") or 0) >= 15:
        reasons.append("Rating excepcional entre los restaurantes del mercado")
    if (prospect.get("reviews_score") or 0) >= 10:
        reasons.append("Alto volumen de resenas — establecimiento consolidado")
    if (prospect.get("zone_score") or 0) >= 10:
        reasons.append("Ubicado en zona de alto potencial comercial")
    if prospect.get("tiene_embutidos"):
        reasons.append("Menu con productos afines a embutidos detectado")
    if not reasons:
        reasons.append("Score compuesto elevado segun analisis ML del sistema")
    return reasons[:3]


def _prospect_card(prospect: dict, rank: int) -> Table:
    RANK_LABELS = ["1 RECOMENDADO", "2 RECOMENDADO", "3 RECOMENDADO"]
    RANK_BG = [colors.HexColor("#FEF3C7"), GRAY_LIGHT_HEX, colors.HexColor("#FEF9C3")]
    RANK_ACCENT = [colors.HexColor("#B45309"), GRAY_HEX, colors.HexColor("#92400E")]

    reasons = _get_reasons(prospect)
    score = prospect.get("total_score", 0)
    card_inner_w = 7.2 * cm

    # Score bar
    score_bar_w = card_inner_w - 0.5 * cm
    bar_draw = Drawing(score_bar_w, 14)
    pct = min(score / 100, 1.0)
    bar_draw.add(Rect(0, 4, score_bar_w, 7, rx=3, ry=3,
                      fillColor=BORDER_HEX, strokeColor=None))
    if pct > 0:
        bar_draw.add(Rect(0, 4, score_bar_w * pct, 7, rx=3, ry=3,
                          fillColor=PRIMARY_HEX, strokeColor=None))

    name_style = ParagraphStyle("cn", fontSize=10, fontName="Helvetica-Bold",
                                textColor=TEXT_DARK, spaceAfter=1)
    meta_style = ParagraphStyle("cm", fontSize=7.5, textColor=GRAY_HEX, spaceAfter=2)
    score_style = ParagraphStyle("cs", fontSize=22, fontName="Helvetica-Bold",
                                 textColor=PRIMARY_HEX, alignment=1)
    score_sub = ParagraphStyle("css", fontSize=7, textColor=GRAY_HEX, alignment=1)
    reason_style = ParagraphStyle("cr", fontSize=7.5, textColor=TEXT_MID, spaceAfter=1)
    rh_style = ParagraphStyle("crh", fontSize=7.5, fontName="Helvetica-Bold",
                               textColor=GRAY_HEX, spaceAfter=3)

    zona = prospect.get("zona") or "Sin zona"
    cocina = prospect.get("tipo_cocina") or "Sin clasificar"
    rating_txt = f"Rating: {prospect['rating']:.1f} / 5" if prospect.get("rating") else "Sin rating"

    card_data = [
        [Paragraph(f"<b>N{rank + 1}  {RANK_LABELS[rank]}</b>",
                   ParagraphStyle("rk", fontSize=8, fontName="Helvetica-Bold",
                                  textColor=RANK_ACCENT[rank], alignment=1))],
        [Paragraph(prospect.get("nombre", ""), name_style)],
        [Paragraph(f"{zona}  /  {cocina}", meta_style)],
        [Paragraph(f"{score:.0f} / 100", score_style)],
        [Paragraph("Score de cliente potencial", score_sub)],
        [bar_draw],
        [Spacer(1, 3)],
        [Paragraph(rating_txt, ParagraphStyle("rt", fontSize=8, textColor=YELLOW_HEX))],
        [Spacer(1, 4)],
        [Paragraph("Por que el sistema lo recomienda:", rh_style)],
        *[[Paragraph(f">  {r}", reason_style)] for r in reasons],
    ]

    t = Table(card_data, colWidths=[card_inner_w])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), RANK_BG[rank]),
        ("BACKGROUND", (0, 1), (-1, -1), colors.white),
        ("BOX", (0, 0), (-1, -1), 1.5, BORDER_HEX),
        ("LINEBELOW", (0, 0), (-1, 0), 1, RANK_ACCENT[rank]),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 3), (-1, 3), "CENTER"),
        ("ALIGN", (0, 4), (-1, 4), "CENTER"),
        ("ALIGN", (0, 5), (-1, 5), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    return t


def _rounded_bar_chart(
    data: list[tuple], title: str,
    width: float = 13 * cm, height: float = 7 * cm,
    bar_color=None,
) -> Drawing:
    from reportlab.graphics.shapes import Line as GLine
    drawing = Drawing(width, height)
    if not data:
        return drawing

    color = bar_color or PRIMARY_HEX
    values = [v for _, v in data]
    labels = [l[:13] for l, _ in data]
    max_val = max(values) * 1.18 if values else 1

    ml, mb, mt, mr = 28, 42, 22, 8
    chart_w = width - ml - mr
    chart_h = height - mb - mt
    n = len(values)
    slot = chart_w / n
    bar_w = slot * 0.62
    gap = (slot - bar_w) / 2

    # Grid lines
    for tick in [0.25, 0.5, 0.75, 1.0]:
        y_g = mb + tick * chart_h
        drawing.add(GLine(ml, y_g, width - mr, y_g,
                          strokeColor=colors.HexColor("#F3F4F6"), strokeWidth=0.5))

    for i, (val, label) in enumerate(zip(values, labels)):
        x = ml + i * slot + gap
        bh = max((val / max_val) * chart_h, 3)

        # Shadow bar (offset 1pt right, lighter)
        drawing.add(Rect(x + 1, mb, bar_w, bh, rx=4, ry=4,
                         fillColor=colors.HexColor("#E5E7EB"), strokeColor=None))
        # Main bar
        drawing.add(Rect(x, mb, bar_w, bh, rx=4, ry=4,
                         fillColor=color, strokeColor=None))
        # Value label
        drawing.add(String(x + bar_w / 2, mb + bh + 2, str(val),
                           fontSize=6.5, textAnchor="middle",
                           fillColor=TEXT_MID))
        # Category label
        drawing.add(String(x + bar_w / 2, mb - 14, label,
                           fontSize=6, textAnchor="middle",
                           fillColor=GRAY_HEX))

    # X axis
    drawing.add(GLine(ml, mb, width - mr, mb,
                      strokeColor=colors.HexColor("#D1D5DB"), strokeWidth=0.8))
    # Title
    drawing.add(String(width / 2, height - 13, title,
                       fontSize=9, fontName="Helvetica-Bold", textAnchor="middle",
                       fillColor=TEXT_DARK))
    return drawing


def _section_header(text: str) -> Paragraph:
    style = ParagraphStyle(
        "sh", fontSize=12, fontName="Helvetica-Bold",
        textColor=PRIMARY_HEX, spaceBefore=4, spaceAfter=6,
        borderPad=0,
    )
    return Paragraph(text, style)


def _kpi_table(kpis: list[tuple]) -> Table:
    bg_colors = [PRIMARY_HEX, GREEN_HEX, colors.HexColor("#7C3AED"), ORANGE_HEX, colors.HexColor("#CA8A04")]
    data = [
        [
            Paragraph(f"<b>{v}</b>",
                      ParagraphStyle("kv", fontSize=18, textColor=colors.white, alignment=1)),
            Paragraph(f"<font size=9>{l}</font>",
                      ParagraphStyle("kl", fontSize=9, textColor=colors.white)),
        ]
        for l, v in kpis
    ]
    style = [("BACKGROUND", (0, i), (-1, i), bg_colors[i % len(bg_colors)]) for i in range(len(data))]
    style += [
        ("BOX", (0, 0), (-1, -1), 0.5, colors.white),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LEFTPADDING", (0, 0), (-1, -1), 12),
    ]
    t = Table(data, colWidths=[3 * cm, 9 * cm])
    t.setStyle(TableStyle(style))
    return t


def export_pdf(db: Session, **filters) -> bytes:
    rows = _get_top_scored(db, limit=50, **filters)
    stats = _get_summary_stats(db)
    prospects = _get_top_prospects_pdf(db, limit=3)

    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output, pagesize=landscape(letter),
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
    )
    elements = []

    title_style = ParagraphStyle("title", fontSize=17, fontName="Helvetica-Bold",
                                 textColor=PRIMARY_HEX, spaceAfter=3)
    subtitle_style = ParagraphStyle("sub", fontSize=9, textColor=GRAY_HEX, spaceAfter=8)

    # === ENCABEZADO ===
    elements.append(Paragraph("Reporte de Inteligencia de Mercado", title_style))
    elements.append(Paragraph(
        f"Fabrica Don Piotr  —  Generado el {datetime.now().strftime('%d de %B de %Y')}",
        subtitle_style,
    ))
    elements.append(HRFlowable(width="100%", thickness=2, color=PRIMARY_HEX, spaceAfter=14))

    # === SECCIÓN 1: RECOMENDACIONES DEL SISTEMA ===
    elements.append(_section_header("Recomendaciones del Sistema  —  Analisis ML"))
    elements.append(Paragraph(
        "Los 3 restaurantes con mayor potencial de conversion segun el modelo de inteligencia "
        "de mercado, excluyendo clientes actuales y prospectos descartados.",
        subtitle_style,
    ))
    elements.append(Spacer(1, 6))

    if prospects:
        cards = [_prospect_card(p, i) for i, p in enumerate(prospects)]
        while len(cards) < 3:
            cards.append(Spacer(1, 1))
        card_col_w = 7.8 * cm
        cards_table = Table(
            [cards],
            colWidths=[card_col_w] * 3,
            hAlign="LEFT",
        )
        cards_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))
        elements.append(cards_table)

    elements.append(Spacer(1, 14))
    elements.append(HRFlowable(width="100%", thickness=0.5, color=BORDER_HEX, spaceAfter=12))

    # === SECCIÓN 2: INDICADORES CLAVE ===
    elements.append(_section_header("Indicadores Clave del Mercado"))
    elements.append(Spacer(1, 4))
    kpis = [
        ("Total Restaurantes", stats["total"]),
        ("Alta Afinidad (Score >= 70)", stats["high_affinity"]),
        ("Clientes Actuales", stats["clients"]),
        ("Con Embutidos en Menu", stats["with_embutidos"]),
        ("Rating Promedio", stats["avg_rating"] or "N/A"),
    ]
    elements.append(_kpi_table(kpis))
    elements.append(Spacer(1, 16))

    # === SECCIÓN 3: GRAFICOS DE DISTRIBUCIÓN ===
    if stats["by_zone"]:
        elements.append(_section_header("Distribucion por Zona y Tipo de Cocina"))
        elements.append(Spacer(1, 4))
        chart_w = 13 * cm
        chart_h = 7 * cm
        zone_chart = _rounded_bar_chart(stats["by_zone"], "Restaurantes por Zona",
                                        chart_w, chart_h, PRIMARY_HEX)
        cuisine_chart = _rounded_bar_chart(stats["by_cuisine"], "Tipos de Cocina mas Frecuentes",
                                           chart_w, chart_h, GREEN_HEX)
        charts_table = Table(
            [[zone_chart, cuisine_chart]],
            colWidths=[chart_w + 1 * cm, chart_w + 1 * cm],
        )
        elements.append(charts_table)
        elements.append(Spacer(1, 16))

    # === SECCIÓN 4: EMBUDO DE VENTAS ===
    if stats["by_status"]:
        elements.append(_section_header("Estado Comercial de Prospectos"))
        elements.append(Spacer(1, 4))
        status_labels = {
            "nuevo": "Nuevo", "contactado": "Contactado", "interesado": "Interesado",
            "cliente": "Cliente", "no_interesado": "No Interesado",
        }
        status_data = [["Estado", "Cantidad"]] + [
            [status_labels.get(k, k), v] for k, v in stats["by_status"].items()
        ]
        st = Table(status_data, colWidths=[6 * cm, 4 * cm])
        st.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), PRIMARY_HEX),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, BORDER_HEX),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [PRIMARY_LIGHT_HEX, colors.white]),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        elements.append(st)
        elements.append(Spacer(1, 20))

    # === SECCIÓN 5: TOP 50 PROSPECTOS ===
    elements.append(HRFlowable(width="100%", thickness=1, color=BORDER_HEX, spaceAfter=10))
    elements.append(_section_header("Top 50 Prospectos Mejor Puntuados"))
    elements.append(Spacer(1, 4))

    pdf_headers = ["#", "Nombre", "Zona", "Cocina", "Rating", "Estado", "Score", "Embutidos"]
    pdf_keys = ["nombre", "zona", "tipo_cocina", "rating", "status", "score", "tiene_embutidos"]

    table_data = [pdf_headers]
    for i, row in enumerate(rows, 1):
        table_data.append([str(i)] + [str(row.get(k, "-"))[:28] for k in pdf_keys])

    col_widths = [1 * cm, 6.5 * cm, 3 * cm, 3.5 * cm, 1.8 * cm, 2.8 * cm, 1.8 * cm, 2.2 * cm]
    prospect_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    prospect_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), PRIMARY_HEX),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("ALIGN", (1, 1), (1, -1), "LEFT"),
        ("ALIGN", (3, 1), (3, -1), "LEFT"),
        ("GRID", (0, 0), (-1, -1), 0.4, BORDER_HEX),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [PRIMARY_LIGHT_HEX, colors.white]),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(prospect_table)

    doc.build(elements)
    return output.getvalue()
